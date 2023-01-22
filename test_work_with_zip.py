import os, csv
from os.path import basename
from zipfile import ZipFile
from PyPDF2 import PdfReader
from io import TextIOWrapper
from openpyxl import load_workbook

path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'to_zip')
path_to = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
path_zip = os.path.join(path_to, "simple-zip.zip")


def test_create_zip():
    file_dir = os.listdir(path)
    with ZipFile(path_zip, "w") as myzip:
        for file in file_dir:
            add_file = os.path.join(path, file)
            myzip.write(add_file, basename(add_file))


def test_read_csv_from_zip():
    zf = ZipFile(path_zip)
    with zf.open("simple-csv.csv") as csvfile:
        csvfile = csv.reader(TextIOWrapper(csvfile))
        list_csv = []
        for r in csvfile:
            text = "".join(r).replace(";", " ")
            list_csv.append(text)
            assert "QuotaAmount StartDate OwnerName Username" in list_csv, f"В файле нет этих данных "
    zf.close()


def test_read_xlsx_from_zip():
    zf = ZipFile(path_zip)
    with zf.open("simple-xls.xlsx") as xlsxfile:
        xlsxfile = load_workbook(xlsxfile)
        sheet = xlsxfile.active
        print(sheet.cell(row=2, column=3).value)
        # assert sheet.cell(row=2, column=3).value == 'Chris Riley'
    zf.close()


def test_read_pdf_from_zip():
    with ZipFile(path_zip) as zf:
        pdf_file = zf.extract("simple-pdf.pdf")
        reader = PdfReader(pdf_file)
        try:
            page = reader.pages[0]
            text = page.extract_text()
            assert "Духовной жаждою томим" in text, \
                f'В файле нет строки Духовной жаждою томим'
        finally:
            os.remove(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'simple-pdf.pdf'))
            zf.close()
